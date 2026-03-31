"""
OKR Tracker Spreadsheet Generator

Usage:
    python scripts/generate_tracker.py --output /path/to/output.xlsx --config /path/to/config.json

The config.json should contain the OKRs in this format:
{
  "company_name": "Acme Corp",
  "quarter": "Q3 2026",
  "date_range": "July 1 - September 30, 2026",
  "objectives": [
    {
      "title": "Prove the product works in live environments",
      "type": "Committed",
      "owner": "Jane Smith",
      "key_results": [
        {
          "description": "3 deployments live and processing by Sept 30",
          "target": "3 deployments",
          "due_date": "2026-09-30",
          "owner": "Jane Smith"
        }
      ]
    }
  ]
}

If no config is provided, generates an empty template.
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule, CellIsRule
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule, CellIsRule


# Style constants
ACCENT = "1A5276"
DARK = "2C3E50"
GREEN = "27AE60"
AMBER = "F39C12"
RED = "E74C3C"
LIGHT_GRAY = "F2F3F4"
WHITE = "FFFFFF"
LIGHT_BLUE = "D4E6F1"
LIGHT_GREEN = "EAFAF1"
LIGHT_AMBER = "FEF5E7"
LIGHT_RED = "FADBD8"

header_font = Font(bold=True, size=11, color=WHITE, name="Arial")
header_fill = PatternFill("solid", fgColor=ACCENT)
subheader_font = Font(bold=True, size=10, color=ACCENT, name="Arial")
body_font = Font(size=10, color=DARK, name="Arial")
bold_font = Font(bold=True, size=10, color=DARK, name="Arial")
title_font = Font(bold=True, size=14, color=ACCENT, name="Arial")
subtitle_font = Font(size=11, color="7F8C8D", name="Arial")

thin_border = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
wrap_top = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")


def style_cell(ws, row, col, value, font=body_font, fill=None, alignment=wrap_top, border=thin_border):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    c.alignment = alignment
    c.border = border
    return c


def create_dashboard(wb, config):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = ACCENT

    # Title
    ws.merge_cells("A1:H1")
    style_cell(ws, 1, 1, f"{config['company_name']} — {config['quarter']} OKR Tracker",
               font=title_font, alignment=Alignment(vertical="center"))
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    style_cell(ws, 2, 1, config.get("date_range", ""), font=subtitle_font)
    ws.row_dimensions[2].height = 22

    # Summary table
    row = 4
    headers = ["#", "Objective", "Type", "Owner", "KRs On Track", "KRs At Risk", "Avg Score", "Status"]
    widths = [4, 45, 14, 18, 14, 14, 12, 12]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        style_cell(ws, row, col, h, font=header_font, fill=header_fill,
                   alignment=Alignment(horizontal="center", vertical="center"))
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, obj in enumerate(config.get("objectives", []), 1):
        row += 1
        n_krs = len(obj.get("key_results", []))
        type_fill = PatternFill("solid", fgColor=LIGHT_GREEN if obj.get("type") == "Committed" else LIGHT_AMBER)

        style_cell(ws, row, 1, i, alignment=center)
        style_cell(ws, row, 2, obj.get("title", f"Objective {i}"))
        style_cell(ws, row, 3, obj.get("type", "Committed"), fill=type_fill, alignment=center)
        style_cell(ws, row, 4, obj.get("owner", ""))
        # KRs on track — counts green from objective sheets
        style_cell(ws, row, 5, "", alignment=center)
        style_cell(ws, row, 6, "", alignment=center)
        style_cell(ws, row, 7, "", alignment=center)
        status_cell = style_cell(ws, row, 8, "Not Started", alignment=center)

    # Add conditional formatting for status column
    last_row = 4 + len(config.get("objectives", []))
    green_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    amber_fill = PatternFill("solid", fgColor=LIGHT_AMBER)
    red_fill = PatternFill("solid", fgColor=LIGHT_RED)

    ws.conditional_formatting.add(f"H5:H{last_row}",
        CellIsRule(operator="equal", formula=['"On Track"'], fill=green_fill))
    ws.conditional_formatting.add(f"H5:H{last_row}",
        CellIsRule(operator="equal", formula=['"Needs Attention"'], fill=amber_fill))
    ws.conditional_formatting.add(f"H5:H{last_row}",
        CellIsRule(operator="equal", formula=['"At Risk"'], fill=red_fill))

    # Instructions
    row = last_row + 3
    style_cell(ws, row, 1, "How to use this tracker:", font=subheader_font)
    row += 1
    instructions = [
        "Each objective has its own tab. Update KR status and scores weekly during check-ins.",
        "Status values: On Track (green), Needs Attention (yellow), At Risk (red), Not Started, Complete.",
        "Confidence: 1-10 scale. How confident are you right now that you'll hit 1.0 by end of quarter?",
        "Score: 0.0-1.0 at end of quarter. For committed OKRs target is 1.0. For aspirational, 0.7 is a good outcome.",
        "Update the dashboard summary after each weekly check-in.",
    ]
    for inst in instructions:
        style_cell(ws, row, 1, inst, font=Font(size=9, color="7F8C8D", name="Arial"))
        ws.merge_cells(f"A{row}:H{row}")
        row += 1

    ws.freeze_panes = "A5"


def create_objective_sheet(wb, obj_num, obj, config):
    title = f"OBJ {obj_num}"
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = GREEN if obj.get("type") == "Committed" else AMBER

    # Header
    ws.merge_cells("A1:I1")
    style_cell(ws, 1, 1, f"OBJECTIVE {obj_num}: {obj.get('title', '')}",
               font=Font(bold=True, size=12, color=ACCENT, name="Arial"),
               alignment=Alignment(vertical="center"))
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    type_label = obj.get("type", "Committed")
    owner_label = obj.get("owner", "TBD")
    style_cell(ws, 2, 1, f"Type: {type_label}  |  Owner: {owner_label}  |  Quarter: {config['quarter']}",
               font=subtitle_font)

    # KR tracking table
    row = 4
    headers = ["KR #", "Key Result", "Target", "Due Date", "Owner", "Status", "Confidence", "Score", "Notes / Blockers"]
    widths = [6, 38, 18, 12, 15, 14, 12, 8, 30]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        style_cell(ws, row, col, h, font=header_font, fill=header_fill,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.column_dimensions[get_column_letter(col)].width = w

    krs = obj.get("key_results", [])
    if not krs:
        # Empty template rows
        krs = [{"description": "", "target": "", "due_date": "", "owner": ""} for _ in range(4)]

    for i, kr in enumerate(krs, 1):
        row += 1
        bg = PatternFill("solid", fgColor=LIGHT_GRAY if i % 2 == 0 else WHITE)
        style_cell(ws, row, 1, f"{obj_num}.{i}", font=bold_font, fill=bg, alignment=center)
        style_cell(ws, row, 2, kr.get("description", ""), fill=bg)
        style_cell(ws, row, 3, kr.get("target", ""), fill=bg)
        style_cell(ws, row, 4, kr.get("due_date", ""), fill=bg, alignment=center)
        style_cell(ws, row, 5, kr.get("owner", obj.get("owner", "")), fill=bg)
        style_cell(ws, row, 6, "Not Started", fill=bg, alignment=center)
        style_cell(ws, row, 7, "", fill=bg, alignment=center)
        style_cell(ws, row, 8, "", fill=bg, alignment=center)
        style_cell(ws, row, 9, "", fill=bg)
        ws.row_dimensions[row].height = 36

    # Status conditional formatting
    last_kr_row = 4 + len(krs)
    green_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    amber_fill = PatternFill("solid", fgColor=LIGHT_AMBER)
    red_fill = PatternFill("solid", fgColor=LIGHT_RED)

    ws.conditional_formatting.add(f"F5:F{last_kr_row}",
        CellIsRule(operator="equal", formula=['"On Track"'], fill=green_fill))
    ws.conditional_formatting.add(f"F5:F{last_kr_row}",
        CellIsRule(operator="equal", formula=['"Needs Attention"'], fill=amber_fill))
    ws.conditional_formatting.add(f"F5:F{last_kr_row}",
        CellIsRule(operator="equal", formula=['"At Risk"'], fill=red_fill))

    # Weekly check-in log
    row = last_kr_row + 3
    ws.merge_cells(f"A{row}:I{row}")
    style_cell(ws, row, 1, "Weekly Check-In Log", font=Font(bold=True, size=11, color=ACCENT, name="Arial"))

    row += 1
    log_headers = ["Week", "Date", "Overall Status", "Key Updates", "Blockers", "Decisions Needed"]
    log_widths_map = {1: 8, 2: 12, 3: 14, 4: 35, 5: 30, 6: 30}
    for col, h in enumerate(log_headers, 1):
        style_cell(ws, row, col, h, font=header_font, fill=PatternFill("solid", fgColor="5D6D7E"),
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

    # Pre-fill 13 weeks
    for week in range(1, 14):
        row += 1
        bg = PatternFill("solid", fgColor=LIGHT_GRAY if week % 2 == 0 else WHITE)
        style_cell(ws, row, 1, f"Wk {week}", font=body_font, fill=bg, alignment=center)
        for col in range(2, 7):
            style_cell(ws, row, col, "", fill=bg)
        ws.row_dimensions[row].height = 28

    # End-of-quarter grading section
    row += 2
    ws.merge_cells(f"A{row}:I{row}")
    style_cell(ws, row, 1, "End-of-Quarter Grading", font=Font(bold=True, size=11, color=ACCENT, name="Arial"))

    row += 1
    grade_headers = ["KR #", "Key Result", "Final Score (0.0-1.0)", "Evidence", "What We Learned"]
    for col, h in enumerate(grade_headers, 1):
        style_cell(ws, row, col, h, font=header_font, fill=PatternFill("solid", fgColor="1ABC9C"),
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

    krs_for_grading = obj.get("key_results", [{"description": ""} for _ in range(4)])
    for i, kr in enumerate(krs_for_grading, 1):
        row += 1
        style_cell(ws, row, 1, f"{obj_num}.{i}", font=bold_font, alignment=center)
        style_cell(ws, row, 2, kr.get("description", ""))
        style_cell(ws, row, 3, "", alignment=center)
        style_cell(ws, row, 4, "")
        style_cell(ws, row, 5, "")
        ws.row_dimensions[row].height = 36

    row += 1
    style_cell(ws, row, 1, "Objective Score:", font=bold_font)
    style_cell(ws, row, 3, "", font=bold_font, alignment=center)

    row += 1
    ws.merge_cells(f"A{row}:I{row}")
    style_cell(ws, row, 1, "Carry forward to next quarter:")

    ws.freeze_panes = "A5"


def generate_tracker(config, output_path):
    wb = openpyxl.Workbook()
    create_dashboard(wb, config)
    for i, obj in enumerate(config.get("objectives", []), 1):
        create_objective_sheet(wb, i, obj, config)
    wb.save(output_path)
    print(f"OKR tracker saved to: {output_path}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Generate OKR tracking spreadsheet")
    parser.add_argument("--output", required=True, help="Output xlsx path")
    parser.add_argument("--config", help="JSON config file with OKR data")
    args = parser.parse_args()

    if args.config:
        with open(args.config) as f:
            config = json.load(f)
    else:
        # Empty template
        config = {
            "company_name": "Company Name",
            "quarter": "Q_ 20__",
            "date_range": "",
            "objectives": [
                {"title": "Objective 1", "type": "Committed", "owner": "", "key_results": []},
                {"title": "Objective 2", "type": "Committed", "owner": "", "key_results": []},
                {"title": "Objective 3", "type": "Aspirational", "owner": "", "key_results": []},
            ]
        }

    generate_tracker(config, args.output)
